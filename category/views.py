from django.shortcuts import render, redirect
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import authentication, permissions
from django.contrib.auth.models import User
from .models import Category, Works, ChangeLog, Specialist
from .serializers import WorkSerializer, CatergorySerializer, SpecialistSerializer
from django.http import HttpResponse, JsonResponse
from legal.settings import BASE_DIR, STATIC_URL
import openpyxl
import pandas as pd
import numpy as np
from fpdf import FPDF
from django.utils import timezone
import random
from string import ascii_uppercase
import json
from io import BytesIO



class CategoryApi(APIView):
    permission_classes = []

    def get(self, request):
        category_list = []
        all_categorys = Category.objects.all()
        # for category in all_categorys:
        #     category_list.append({ 'name': category.name, 'id': category.id })
        # categorys = { 'categories': category_list, }
        ser = CatergorySerializer(all_categorys, many=True)
        categorys = {'categories': ser.data, }
        return Response(categorys)

class WorkApi(APIView):
    permission_classes = []

    def get(self, request):
        """
        параметр category -- айди категории

        """
        #category_id  =request.GET.get('category')
        category_list = []
        all_works = Works.objects.all()
        ser = WorkSerializer(all_works, many=True)
        # for category in all_categorys:
        #     category_list.append({ 'name': category.name, 'id': category.id })
        categorys = { 'services': ser.data, }
        return JsonResponse(categorys, safe=True)

class ChangeLogApi(APIView):
    permission_classes = []

    def get(self, request):
        log = ChangeLog.objects.all()
        if not log:
            ChangeLog.objects.create()
        else:
            log = log[0]
        logs = {
            'change': log.value
        }
        return JsonResponse(logs, safe=True)

class SpecialistsApi(APIView):
    permission_classes = []

    def get(self, request):
        specialist = Specialist.objects.all()
        ser = SpecialistSerializer(specialist, many=True)
        categorys = {'specialists': ser.data, }
        return Response(categorys)


class PDFApi(APIView):
    permission_classes = []

    def post(self, request):
        print(request.POST, 'это реквест')
        # print(request.POST.getlist('works'))
        try:
            b = json.loads(list(request.POST.keys())[0])
            req = b['works']
        #b = request.POST
        except:
            b = request.POST
            req = json.loads(b['works'])
        print(b, 'распаршен')
        excel = create_excel(req)
        print(excel, 'это эксель')
        client_name = b['client_name']
        client_itn = b['client_itn']
        final_price = b['final_price']
        bd_receiver = b['bd_receiver']
        bd_bank_receiver = b['bd_bank_receiver']
        bd_tin = b['bd_tin']
        bd_iec = b['bd_iec']
        bd_purpose_of_payment = b['bd_purpose_of_payment']
        bd_bic = b['bd_bic']
        bd_check_account = b['bd_check_account']
        bd_cor_account = b['bd_cor_account']
        bd_signing = b['signing_person']
        is_discount = b['is_discount']
        is_fixed = b['is_fixed']
        payer_info = b['payer_info']
        recalculation_value = b['recalculation_value']
        pdf = create_pdf(excel, client_name, client_itn, final_price, bd_receiver, bd_bank_receiver, bd_tin, bd_iec, bd_purpose_of_payment, bd_bic, bd_check_account, bd_cor_account, bd_signing, is_discount, is_fixed, recalculation_value, payer_info)
        print(pdf)
        host = request.get_host()
        if request.is_secure():
            return Response({'pdf': f'https://{host}{pdf}'})
        else:
            return Response({'pdf': f'http://{host}{pdf}'})

# Create your views here.


def create_excel(works):
    random_file_name = 'doc_'
    random_file_name += str(random.randint(0, 100000))
    random_file_name += '_' + str(timezone.now().date()).replace('.', '_').replace(':', '_') + '_'
    random_file_name += str(random.randint(0, 100000)) + '.xlsx'
    file_path = BASE_DIR + STATIC_URL + random_file_name
    path = BASE_DIR + STATIC_URL + 'test.xlsx'
    arr = []
    book = openpyxl.load_workbook(path)
    ws = book.worksheets[0]
    self_json = {"works":
                     [
                         {'name':"Банкротство. Сопровождение процесса по кредиторам/по привлечению к субсидиарной ответственности",
                          'specialistID':1,"number":1,
                          "hours":1,"price":1523,"count":1},
                          {"price":1523,"number":2,"count":1,"specialistID":1,'name':"По dfdfdf fdfdfdfd fddfdf fdfdfd dfd d","hours":1},
                         {"hours":1,"count":4,"price":1523,"number":3,'name':"По","specialistID":1},
                         {"specialistID":1,"count":1,"price":1523,"hours":7,"number":4,'name':"По"},
                         {"count":1,"price":1523,"number":5,"specialistID":1,'name':"Орга","hours":1},
                         {"price":1523,'name':"Сопровождение купли-продажи доли в ООО","count":1,"hours":1,"specialistID":1,"number":6}
                     ]
    }

    ws['A1'] = '№'
    ws['B1'] = 'Вид работ'
    ws['C1'] = 'Тип специалиста'
    ws['D1'] = 'Часы'
    ws['E1'] = 'Количество'
    ws['F1'] = 'Стоимость'
    i = 0
    print(works, 'это длинна')
    #print(works['works'], 'sdf')
    #test_json = json.loads(works['works'])
    #print(test_json, 'asdad')
    for row in range(2, len(works) + 2):
        #print(works[row])
        specialist = Specialist.objects.get(id=int(works[i]['specialistID']))
        ws['A' + str(row)] = works[i]['number']
        ws['B' + str(row)] = works[i]['name']
        price = round(works[i]['price'], 2)
        ws['F' + str(row)] = price
        #print(works['works'][i]['hours'])
        #hours = round(works[i]['hours'], 2)
        ws['D' + str(row)] = works[i]['hours']

        ws['E' + str(row)] = works[i]['count']
        ws['C' + str(row)] = specialist.name
        i = i + 1
    #print(works)
    file_name = str('rtty') + str('sdf') + '.xlsx'
    file = BytesIO()
    book.save(file)
    file.seek(0)
    return file

def create_pdf(excel, name, inn, final_price, bd_receiver, bd_bank_receiver, bd_tin, bd_iec, bd_purpose_of_payment, bd_bic, bd_check_account, bd_cor_account, bd_signing, is_discount, is_fixed, recalculation_value, payer_info):
    random_file_name = 'doc_'
    random_file_name += str(random.randint(0, 100000))
    random_file_name += '_' + str(timezone.now().date()).replace('.', '_').replace(':', '_') + '_'
    random_file_name += str(random.randint(0, 100000)) + '.pdf'
    #df_1 = pd.DataFrame(np.random.randn(10, 2), columns=list('AB'))
    
    #df_1.to_excel(writer)

    # read in the .xlsx file just created
    df_2 = pd.read_excel(excel, engine="openpyxl")
    writer = pd.ExcelWriter(excel)
    # creating a pdf in called test.pdf in the current directory
    pdf = FPDF()
    # Save top coordinate


# Calculate x position of next cell

    pdf.add_page()
    pdf.set_xy(0, 0)

    #pdf.set_doc_option("core_fonts_encoding", 'cp437')
    #pdf.set_font('arial', 'B', 14)
    pdf.add_font('DejaVu', '', 'DejaVuSansCondensed.ttf', uni=True)
    pdf.set_font('DejaVu', '', 10)
    pdf.cell(20)
    #currents_y = pdf.y
    currents_x = pdf.x
    inn_x = pdf.x + 70
    inn_y = pdf.y
    pdf.multi_cell(70, 10, 'Клиент:', 0, 0, 'L')
    pdf.y = inn_y
    pdf.x = inn_x
    pdf.cell(70, 10, 'Исполнитель:', 0, 2, 'L')

    #pdf.y = currents_y
    pdf.x = currents_x
    pdf.cell(70, 5, name, 0, 0, 'L')
    pdf.cell(70, 5, 'Получатель: ' + bd_receiver, 0, 2, 'L')

    #pdf.y = currents_y
    pdf.x = currents_x
    pdf.cell(70, 5, inn, 0, 0, 'L')
    pdf.cell(70, 5, 'ИНН:' + ' ' + bd_tin, 0, 2, 'L')
    pdf.cell(70, 5, 'КПП:' + ' ' + bd_iec, 0, 2, 'L') if bd_iec != '' else print()
    pdf.cell(70, 5, 'Банк получателя:' + ' ' + bd_bank_receiver, 0, 2, 'L') if bd_bank_receiver != '' else print()

    #pdf.y = currents_y
    #pdf.x = currents_x
    pdf.cell(70, 5, 'K/c:' + ' ' + bd_cor_account, 0, 2, 'L') if bd_cor_account != '' else print()
    pdf.cell(70, 5, 'БИК:' + ' ' + bd_bic, 0, 2, 'L') if bd_bic != '' else print()

    #pdf.y = currents_y
    #pdf.x = currents_x
    pdf.cell(70, 5, 'P/c:' + ' ' + bd_check_account, 0, 2, 'L') if bd_check_account != '' else print()
    #pdf.x = currents_x
    pdf.multi_cell(140, 5, 'Назначение платежа:' + ' ' + bd_purpose_of_payment, 0, 2, 'L') if bd_purpose_of_payment !=  '' else print()
    #pdf.cell(0)
    pdf.x = currents_x
    pdf.cell(5, 10, '№', 1, 0, 'C')
    pdf.cell(60, 10, 'Виды работ', 1, 0, 'C')
    top = pdf.y
    offset = pdf.x + 28
    pdf.multi_cell(28, 5, 'Тип\nспециалиста', 1, 0, 'C')
    pdf.y = top
    pdf.x = offset
    pdf.cell(20, 10, 'Часы', 1, 0, 'C')
    pdf.cell(23, 10, 'Количество', 1, 0, 'C')
    pdf.cell(20, 10, 'Стоимость', 1, 2, 'C')
    pdf.cell(-136)
    #pdf.set_font('arial', '', 12)
    # print(type(df_2))
    begin_y = pdf.y
    end_y = 0
    number_y = pdf.y
    number_x = pdf.x
    for i in range(0, len(df_2)):
        pdf.x += 5
        col_ind = str(i)
        # print(df_2)
        col_a = str(df_2['№'].iloc[i])
        col_b = str(df_2['Вид работ'].iloc[i])
        col_с = str(df_2['Тип специалиста'].iloc[i])
        col_e = str(df_2['Часы'].iloc[i])
        col_f = str(df_2['Количество'].iloc[i])
        col_g = str(df_2['Стоимость'].iloc[i])
        #pdf.cell(15, 10, '%s' % (col_ind), 1, 0, 'C')
        #pdf.cell(5, 10, '%s' % (col_a), 1, 0, 'C')
        top = pdf.y
        offset = pdf.x + 60

        pdf.multi_cell(60, 5, '%s' % (col_b), 1, 0, 'L')
        test_y = pdf.y
        pdf.y = top
        pdf.x = offset
        return_x = pdf.x
        return_y = pdf.y
        pdf.x = number_x
        pdf_y = number_y


        # offset = pdf.x + 35
        current_y = test_y - begin_y - end_y
        pdf.cell(5, current_y, '%s' % (col_a), 1, 0, 'C')
        pdf.x = return_x
        pdf_y = return_y

        if len(col_с) > 8:
            height = current_y / 2
        else:
            height = current_y
        # print(col_с)
        if col_с == 'Технический специалист':
            print(True)
        c_y = pdf_y
        c_x = pdf.x + 28
        pdf.multi_cell(28, height, '%s' % (col_с), 1, 0, 'C')
        pdf.y = c_y
        pdf.x = c_x
        # print()
        # pdf.y = top
        # pdf.x = offset
        pdf.cell(20, current_y, '%s' % (col_e), 1, 0, 'C')
        pdf.cell(23, current_y, '%s' % (col_f), 1, 0, 'C')
        pdf.cell(20, current_y, '%s' % (col_g), 1, 1, 'C')
        pdf.cell(10)
        end_y += current_y
        number_y = pdf.y
        number_x = pdf.x
        # print(pdf.y)
    if is_discount == True or is_discount == 'true':
        first_value = 'Скидка'
    else:
        first_value = 'Наценка за сложность'
    if is_fixed == True or is_fixed == 'true':
        fix = 'р.'
    else:
        fix = '%'
    pdf.cell(134, 5, first_value, 1, 0, 'L')
    pdf.cell(22, 5, str(recalculation_value)+fix, 1, 1, 'L')
    pdf.x = currents_x
    pdf.cell(134, 5, 'Итого', 1, 0, 'L')
    pdf.cell(22, 5, final_price+'р.', 1, 2, 'L')
    pdf.x = currents_x
    pdf.cell(70, 5, payer_info, 0, 2, 'L')
    pdf.cell(70, 5, bd_signing, 0, 2, 'L')
    df_2.to_excel(writer)
    writer._save()
    pdf.output(BASE_DIR + STATIC_URL + random_file_name, 'F')
    return(STATIC_URL + random_file_name)

def parce_excel(request):
    path = BASE_DIR + STATIC_URL + 'Виды работ.xlsx'
    print('первый уровень')
    works = []
    groups = []
    doc = openpyxl.load_workbook(path)
    if len(doc.worksheets) == 0:
        print('no worksheets')
    sheet = doc.worksheets[0]
    print(sheet)
    cols = list(ascii_uppercase)
    end_row = 186
    print(end_row, 'энд роу')
    have_dates = False
    num_col = 'A'  # Столбец номера или наименования раздела
    name_col = 'B'  # Столбец для проверки по ГЭСН
    detail_col = 'C'
    #name_col = 'C'  # Столбец наименования
    count_type_col = 'D'  # Столбец единиц измерения
    hours_col = 'E'  # Столбец количества
    price_col = 'F'  # Столбец общей стоимости
    count_col = 'F'
    start_date_col = 'H'
    print('второй уровень')
    # start_date_col_num = 4
    # end_date_col = 'H'
    # end_date_col_num = 4
    # date_cols = {}  # Масив колонок плановых дат выполнения работы
    columns_count = 0
    for col in sheet.columns:
        columns_count += 1
    print(columns_count)
    # Найти шапку Таблицы
    row = 1
    print(sheet['B4'].value)
    while 'Категория' not in str(sheet[name_col + str(row)].value):
        row += 1
        #print(str(sheet[name_col + str(row)].value).lower())
    # Найти первый раздел
    # while sheet[num_col + str(row)].value is None or sheet[num_col + str(row)].value != 1 \
    #         and row < end_row:
    #     print(sheet[num_col + str(row)].value)
    #     row += 1
    start_row = row
    print(row, 'ров')
    print('treriy uroven')
    print('start_row', start_row)

    #print('date_cols', date_cols)
    print('Текущая строка после обработки шапки: ' + str(row))
    print('На первом разделе')
    print('pytuy')
    # Ожидание, что нет пустых ячеек в колонке А в теле таблицы
    while row <= end_row and sheet[name_col + str(row)].value != None:
        #print(str(sheet[reason_col + str(row)].value), 'СМОТРИ СЮДА')
        #print('row:', row)
        name = sheet[name_col + str(row)].value
        name_or_num = sheet[num_col + str(row)].value
        try:
            name_or_num = int(name_or_num)
            building = True
        except:
            building = False
            if 'Категория' not in sheet[name_col + str(row)].value:
                dopoltinelno = True
            else:
                dopoltinelno =False
        #pattern = re.compile('раздел \d', re.IGNORECASE)
        #match = re.findall(pattern, name_or_num)
        if building == True:
            # Это здание
            # name = name_or_num.replace(match[0] + '.', '').strip()
            # num = int(match[0].lower().replace('раздел', ''))
            work_name = sheet[name_col + str(row)].value
            works.append({
                'name': name,
                'num': name_or_num
            })
            work = Works.objects.create(number=sheet[num_col + str(row)].value, name=work_name, hours=sheet[hours_col + str(row)].value, count=sheet[count_col + str(row)].value if sheet[count_col + str(row)].value is not None else 1, )
            category.services.add(work)
        else:
            name = sheet[name_col + str(row)].value
            if dopoltinelno == True:
                category = Category.objects.create(name='Дополнительно - ' + category.name, isOriginal=False)
            else:
                category = Category.objects.create(name=name, isOriginal = True)
        row += 1
    #print('works', works)
    return HttpResponse('true')
